import pandas as pd
import os
import chardet
from openpyxl.styles import PatternFill
from datetime import datetime

start_time = datetime.now()


def load_commission_rates():
    """Загружает ставки комиссий из файла setup.xlsx"""
    default_rates = {
        'MIR': 0.0142,
        'MC': 0.02,
        'VISA': 0.0165,
        'CUP': 0.0165,
        'AMEX': 0.03,  # Американ Экспресс
        'DEFAULT': 0.00  # Используем ставку MIR по умолчанию
    }

    try:
        setup_path = os.path.join(os.getcwd(), 'setup.xlsx')
        if not os.path.exists(setup_path):
            print("Файл setup.xlsx не найден, используются ставки по умолчанию")
            return default_rates

        df = pd.read_excel(setup_path)
        if df.empty:
            print("Файл setup.xlsx пуст, используются ставки по умолчанию")
            return default_rates

        commission_rates = {}
        for _, row in df.iterrows():
            card_type = str(row['Тип карты']).strip().upper()
            rate = float(row['Ставка комиссии'])
            commission_rates[card_type] = rate

        commission_rates.setdefault('DEFAULT', default_rates['DEFAULT'])
        return commission_rates

    except Exception as e:
        print(f"Ошибка при загрузке ставок комиссий: {str(e)}, используются ставки по умолчанию")
        return default_rates


def detect_encoding(file_path):
    """Определяет кодировку файла"""
    with open(file_path, 'rb') as f:
        raw_data = f.read(10000)
    return chardet.detect(raw_data)['encoding']


def read_file_with_encoding(file_path):
    """Читает файл с автоматическим определением кодировки и разделителя"""
    try:
        if file_path.endswith(('.csv', '.dsv', '.dsvp')):
            # Пробуем UTF-8 с разделителем ;
            try:
                df = pd.read_csv(file_path, encoding='utf-8', delimiter=';', decimal=',')
                return df, 'utf-8'
            except UnicodeDecodeError:
                try:
                    # Пробуем Windows-1251 с разделителем ;
                    df = pd.read_csv(file_path, encoding='windows-1251', delimiter=';', decimal=',')
                    return df, 'windows-1251'
                except UnicodeDecodeError:
                    # Пробуем определить кодировку автоматически
                    encoding = detect_encoding(file_path)
                    df = pd.read_csv(file_path, encoding=encoding, delimiter=';', decimal=',')
                    return df, encoding
        else:
            df = pd.read_excel(file_path)
            return df, None
    except Exception as e:
        raise ValueError(f"Ошибка чтения файла {file_path}: {str(e)}")


def convert_amount(value):
    """Конвертирует сумму в float, обрабатывая разные форматы"""
    if isinstance(value, str):
        # Удаляем пробелы как разделители тысяч и заменяем запятую на точку
        value = value.replace(' ', '').replace(',', '.')
    return float(value)


def calculate_commission(row, commission_rates):
    """Рассчитывает комиссию с проверкой типа операции"""
    if row['TYPE'] != 'ПОКУПКА':
        return 0.0

    try:
        card_type = str(row['PMT_SYSTEM_CODE']).strip().upper()
        amount = convert_amount(row['AMOUNT'])
        rate = commission_rates.get(card_type, commission_rates['DEFAULT'])
        commission = round(amount * rate, 2)
        return commission
    except Exception as e:
        print(f"Ошибка расчета комиссии для строки: {row}. Ошибка: {str(e)}")
        return 0.0


def process_file(file_path, results_df, commission_rates):
    """Обрабатывает один файл и возвращает обновленный DataFrame с расхождениями"""
    try:
        print(f"\nОбработка файла: {os.path.basename(file_path)}")
        df, original_encoding = read_file_with_encoding(file_path)

        # Проверка необходимых колонок
        required_columns = ['TYPE', 'AMOUNT', 'COMMISSION', 'PMT_SYSTEM_CODE']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            raise ValueError(f"Отсутствуют обязательные колонки: {', '.join(missing_columns)}")

        # Конвертируем числовые колонки
        df['AMOUNT'] = df['AMOUNT'].apply(convert_amount)
        df['COMMISSION'] = df['COMMISSION'].apply(convert_amount)

        # Добавляем расчетные колонки
        df['Комиссия (расчет)'] = df.apply(lambda row: calculate_commission(row, commission_rates), axis=1)
        df['Разница (F - U)'] = df.apply(
            lambda row: (row['COMMISSION'] - row['Комиссия (расчет)']) if row['TYPE'] == 'ПОКУПКА' else 0.0,
            axis=1
        )
        df['Разница (F - U)'] = df['Разница (F - U)'].round(2)

        # Сохраняем ВСЕ расхождения (отличные от нуля)
        discrepancies = df[df['Разница (F - U)'] != 0].copy()
        if not discrepancies.empty:
            discrepancies['Файл'] = os.path.basename(file_path)
            results_df = pd.concat([results_df, discrepancies], ignore_index=True)

        # Сохраняем файл в XLSX
        base_name = os.path.splitext(os.path.basename(file_path))[0]
        new_path = os.path.join(os.path.dirname(file_path), f"{base_name}_processed.xlsx")

        with pd.ExcelWriter(new_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False)
            wb = writer.book
            ws = wb.active

            # Окрашивание
            u_col = df.columns.get_loc('Комиссия (расчет)') + 1
            v_col = df.columns.get_loc('Разница (F - U)') + 1

            green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
            red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            gray_fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=u_col, max_col=v_col):
                if row[0].value is not None:  # Колонка U
                    row[0].fill = green_fill if row[0].value != 0 else gray_fill

                if row[1].value is not None:  # Колонка V
                    if row[1].value == 0:  # Нулевая разница
                        row[1].fill = gray_fill
                    elif row[1].value < 0:  # Отрицательная разница
                        row[1].fill = red_fill
                    else:  # Положительная разница
                        row[1].fill = green_fill

        print(f"Файл обработан и сохранен как: {new_path}")
        return results_df

    except Exception as e:
        print(f"Ошибка при обработке файла {file_path}: {str(e)}")
        return results_df


def main():
    """Основная функция программы"""
    # Загружаем ставки комиссий
    commission_rates = load_commission_rates()
    print("Используемые ставки комиссий:", commission_rates)

    # Создаем пустой DataFrame для результатов
    results_df = pd.DataFrame()

    # Получаем список файлов для обработки
    current_dir = os.getcwd()
    processed_files = [
        os.path.join(current_dir, f) for f in os.listdir(current_dir)
        if (f.endswith('.xlsx') or f.endswith('.xls') or
            f.endswith('.csv') or f.endswith('.dsv') or f.endswith('.dsvp'))
           and not f.startswith('results')
           and not f.endswith('_processed.xlsx')
           and not f.endswith('_processed.xls')
           and not f.endswith('_processed.csv')
           and f != 'setup.xlsx'
    ]

    if not processed_files:
        print("Не найдено файлов для обработки")
        return results_df

    # Обрабатываем файлы
    for file_path in processed_files:
        results_df = process_file(file_path, results_df, commission_rates)

    # Сохраняем результаты
    results_path = os.path.join(current_dir, 'results.xlsx')
    if os.path.exists(results_path):
        open(results_path, 'w').close()

    if not results_df.empty:
        results_df['Время обработки'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        with pd.ExcelWriter(results_path, engine='openpyxl') as writer:
            results_df.to_excel(writer, index=False)
            wb = writer.book
            ws = wb.active

            red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
            gray_fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')

            diff_col = results_df.columns.get_loc('Разница (F - U)') + 1
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=diff_col, max_col=diff_col):
                if row[0].value is not None:
                    if row[0].value == 0:
                        row[0].fill = gray_fill
                    elif row[0].value < 0:
                        row[0].fill = red_fill
                    else:
                        row[0].fill = green_fill

        print(f"\nСводный файл с расхождениями сохранен как: {results_path}")
        print(f"Всего найдено расхождений (≠0): {len(results_df)}")
    else:
        print("\nРасхождений (≠0) не обнаружено")
    print(f'Программа завершилась за {datetime.now() - start_time}')
    print('Нажмите на ENTER для продолжения')
    input()
    return results_df


if __name__ == "__main__":
    results = main()